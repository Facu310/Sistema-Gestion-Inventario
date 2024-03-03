import openpyxl
from PyQt6 import QtWidgets
import re
from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl.styles import PatternFill

from bd_conexion import GestionInventario


class MisMetodos:
    list_prods_tab4=list()
    last_id_sale=str()
    @staticmethod
    def validar_producto(nombre=None, precio=None, stock=None, cat=None, descripcion=None):

        if len(nombre) < 4:
            return "Debe ingresar un nombre de producto adecuado"
        elif not MisMetodos.es_float(precio) or float(precio) <= 0:
            return "Debe ingresar un precio válido"
        elif  MisMetodos.es_float(cat) or len(cat) < 4:
            return "debe ingresar una categoria valida"
        elif not stock.isdigit() or int(stock) <= 0:
            return "Debe ingresar un stock válido"
        elif len(descripcion) <= 0:
            return "Igrese una descripcion"
        else:
            return "Todos los productos validados correctamente"

    @staticmethod
    def validate_one_data_prod(data:str,column_name:str):
        if column_name=="Nombre":
            if len(data) < 4:
                return "Debe ingresar un nombre de producto adecuado"
        if column_name == "Descripción":
            if len(data) <= 0:
                return "Igrese una descripcion"
        if column_name == "Categoria":
            if MisMetodos.es_float(data) or len(data) < 4:
                return "debe ingresar una categoria valida"
        if column_name =="Precio":
            if not MisMetodos.es_float(data) or float(data) <= 0:
                return "Debe ingresar un precio válido"
        if column_name=="Stock":
            if not data.isdigit() or int(data) <= 0:
                return "Debe ingresar un stock válido"
        else:
            return "Atributo validado correctamente"



    @staticmethod
    def es_float(valor):
        try:
            float(valor)
            return True
        except ValueError:
            return False

    @staticmethod
    def es_int(valor):
        try:
            int(valor)
            return True
        except ValueError:
            return False

    @staticmethod
    def write_table_prod(Qlist,data):

       try:
        for i in data:
            item_text = f"{i[0]},{i[1]}"
            item = QtWidgets.QListWidgetItem(item_text)
            Qlist.addItem(item)
       except Exception as e:
           print("An error has occurred", e)

    @staticmethod
    def serch_supplier(Qlist,data):
       try:
        for i in data:
            item_text = f"{i[0]},{i[1]},{i[2]}"
            item = QtWidgets.QListWidgetItem(item_text)
            Qlist.addItem(item)
       except Exception as e:
           print("An error has occurred", e)

    @staticmethod
    def validate_supplier(rut=None,nom=None, contacto=None, termino_pago=None):
        if rut==None or not MisMetodos.validate_rut(rut):
            return "¡Debe ingresar un rut correcto!"
        elif len(nom) <= 0 or nom==None:
            return "Debe ingresar un nombre al proveedor"
        elif contacto==None:
            return "Debe ingresar un número de teléfono de contacto"
        else:
            # Expresión regular para validar números de teléfono de Uruguay
            patron_uruguay = r'^(\+598|0)9\d{7}$'

            if not re.match(patron_uruguay, contacto):
                return "Número de teléfono no válido para Uruguay"

        if len(termino_pago) <= 0 or termino_pago==None:
            return "Ingrese los términos de pago"
        else:
            return"provedor agregado correctamente"
    @staticmethod
    def validar_float(dato):
        return re.match(r'^\d+\.\d+$', dato) is not None
    @staticmethod
    def validate_rut(rut):
        patron = r'^\d{2}\.\d{3}\.\d{3}-\d$'
        return re.match(patron, rut) is not None
    @staticmethod
    def get_prod_link_supplier(selected_items):
        # almacenara los id, nombre en tuplas de los productos
        products = []
        if selected_items:
            for item in selected_items:
                texto = item.text()
                id, nombre = texto.split(',')
                id = int(id)
                products.append((id, nombre))
        return products

    @staticmethod
    def validate_orders(id_suppier,date_initial,date_delivery,price,state,cantidad,description,list_prod:list):
       try:
        date_initial_datetime = datetime.combine(date_initial, datetime.min.time())
        date_delivery_datetime= datetime.combine(date_delivery, datetime.min.time())
        date_year_before = datetime.now() - relativedelta(years=1)
        if not GestionInventario.view_supier(id_suppier):
            return "Ese proveedor no existe,agregelo antes de hacer un pedido."
        elif date_initial_datetime <= date_year_before:
            return "ingrese una fecha adecuada entre un año de antiguedad."
        elif date_delivery_datetime < date_initial_datetime:
            return "la fehca de entrega no puede ser menor a la de el pedido."
        elif price=="" or float(price)<0:
            return "ingrese un precio correcto."
        elif state not in ["Entregado", "Pendiente", "Cancelado"]:
            return "ingrese un estado del preducto"
        elif not MisMetodos.es_int(cantidad) or int(cantidad)<1:
            return "ingrese una cantidad valida."
        elif not list_prod:
            return "Seleccione al menos un producto"
        else:
            GestionInventario.add_order(id_suppier, date_initial_datetime, date_delivery_datetime, price, state,description)
            id_order=GestionInventario.get_last_id("OrdenDeCompra")
            print("la id de la orden es: ",id_order)
            for prod in list_prod:
                GestionInventario.add_order_prod(id_order,prod[0],cantidad)
            return "orden ingresada correctamente"
       except Exception as e:
           print(e)

    @staticmethod
    def validate_cahnge_atributes_order(date_initial,date_delivery,price,state,cantidad,description):
        try:
            date_initial_datetime = datetime.combine(date_initial, datetime.min.time())
            date_delivery_datetime = datetime.combine(date_delivery, datetime.min.time())
            date_year_before = datetime.now() - relativedelta(years=1)


            if date_initial_datetime <= date_year_before:
                return "ingrese una fecha adecuada entre un año de antiguedad."
            elif date_delivery_datetime < date_initial_datetime:
                return "la fehca de entrega no puede ser menor a la de el pedido."
            elif price == "" or float(price) < 0:
                return "ingrese un precio correcto."
            elif state not in ["Entregado", "Pendiente", "Cancelado"]:
                return "ingrese un estado del preducto"
            elif not MisMetodos.es_int(cantidad) or int(cantidad) < 1:
                return "ingrese una cantidad valida."
        except Exception as e:
            print(e)

    @staticmethod
    def validate_update_order(column,new_value,id_order,value_2=None):
       try:
        if column == "Fecha Realiza":
            column="fecha_realiza"
            fecha_object1=datetime.strptime(new_value, "%Y-%m-%d").date()
            fecha_object2=datetime.strptime(value_2, "%Y-%m-%d").date()
            date_initial_datetime = datetime.combine(fecha_object1, datetime.min.time())
            date_delivery_datetime = datetime.combine(fecha_object2, datetime.min.time())
            date_year_before = datetime.now() - relativedelta(years=1)
            if date_initial_datetime <= date_year_before:
                return "ingrese una fecha adecuada entre un año de antiguedad."
            elif date_delivery_datetime < date_initial_datetime:
                return "la fehca de entrega no puede ser menor a la de el pedido."
            else:
                GestionInventario.uptade_orders(column, new_value, id_order)
                return "actualizado correctamente"
        elif column == "Fecha Entrega":
            column="fecha_entrega"
            fecha_object1 = datetime.strptime(new_value, "%Y-%m-%d").date()
            fecha_object2 = datetime.strptime(value_2, "%Y-%m-%d").date()
            date_initial_datetime = datetime.combine(fecha_object2, datetime.min.time())
            date_delivery_datetime = datetime.combine(fecha_object1, datetime.min.time())
            date_year_before = datetime.now() - relativedelta(years=1)
            if date_initial_datetime <= date_year_before:
                return "ingrese una fecha adecuada entre un año de antiguedad."
            elif date_delivery_datetime < date_initial_datetime:
                return "la fehca de entrega no puede ser menor a la de el pedido."
            else:
                GestionInventario.uptade_orders(column, new_value, id_order)
                return "actualizado correctamente"
        elif column == "Precio":
            column="precio"
            if not MisMetodos.validar_float(new_value) or float(new_value) < 0:
                return "ingrese un precio correcto."
            else:
                GestionInventario.uptade_orders(column, new_value, id_order)
                return "actualizado correctamente"
        elif column == "Estado":
            column = "estado"
            if new_value not in ["Entregado", "Pendiente", "Cancelado"]:
                return "ingrese un estado correcto del producto: Entregado ,Pendiente ,Cancelado"
            else:
                GestionInventario.uptade_orders(column, new_value, id_order)
                return "actualizado correctamente"
        elif column == "Cantidad":
            column = "cantidad"
            if not MisMetodos.es_int(new_value) or int(new_value) < 1:
                return "ingrese una cantidad valida."
            else:
               GestionInventario.uptade_orders_prod(column,new_value,id_order,value_2)
               return "actualizado correctamente"

       except Exception as e:
           print(e)

    @staticmethod
    def validate_order_baja(id_order,payment_date):

        date = datetime.combine(payment_date, datetime.min.time())
        date_year_before = datetime.now() - relativedelta(years=1)
        date_year_after = datetime.now() + relativedelta(years=1)
        print("date es:",date)
        print("date_year_after",date_year_after)
        print("date_year_before",date_year_before)

        try:
            if not GestionInventario.exist_id_order(id_order):
                return "Ese id no existe"
            elif  not(date_year_before <= date and date <= date_year_after):
                return "la fecha debe ser entre 1 año"
            else:
                GestionInventario.add_orders_placed(id_order,payment_date)
                return "dado de baja correctamente"

        except Exception as e:
            print(e)


    @staticmethod
    def validate_sales_record(sale_date:datetime,total_amount:str):
       try:
        date = datetime.combine(sale_date, datetime.min.time())
        date_year_before = datetime.now() - relativedelta(years=1)
        date_year_after = datetime.now() + relativedelta(years=1)
        if not (date_year_before <= date and date <= date_year_after):
            return "la fecha debe ser entre 1 año"
        elif not MisMetodos.es_float(total_amount) or float(total_amount)<1:
            return "Ingrese un precio valido"
        else:
            GestionInventario.add_sales_record(sale_date,total_amount)
            MisMetodos.last_id_sale=GestionInventario.get_last_id("Ventas")
            print("atributo MisMetodos.last_id_sale linea 266:",MisMetodos.last_id_sale)
            return "Venta agregada exitosamente."
       except Exception as e:
           print(e)

    @staticmethod
    def on_selection_changed(selected, model):
        indexes = selected.indexes()
        if indexes:
            row = indexes[0].row()
            first_name_item = model.item(row, 0)
            last_name_item = model.item(row, 1)
            if first_name_item and last_name_item:
                first_name = first_name_item.text()
                last_name = last_name_item.text()
                data=(first_name,last_name)
                if data in MisMetodos.list_prods_tab4:
                    print("este producto ya se ha seleccionado")
                else:
                    MisMetodos.list_prods_tab4.append((first_name,last_name))
        else:
            print("Ingrese un producto")


    @staticmethod
    def validate_unit_priece(VentaId,ProductoId,cant,priece):
        index=0
        lsit_aux=MisMetodos.list_prods_tab4
        if not MisMetodos.es_float(priece) or float(priece)<=0:
            return "Ingrese un precio valido y una cantidad valida"
        elif not MisMetodos.es_int(cant) or float(cant) <= 0:
            return "Ingrese una cantidad valida"
        else:
            print("atrubito VentaId:",VentaId)
            print("ATRUBITO ProductoId",ProductoId)
            print("atributo cant",cant)
            print("atributo priece:",priece)
            GestionInventario.add_sales_record_prod(VentaId,ProductoId,cant,priece)
            print("lista de productos list_prods_tab4 antes de eliminar ",MisMetodos.list_prods_tab4)
            if index < len(lsit_aux):
                print(index)
                MisMetodos.list_prods_tab4.pop(index)
                print("lista de productos list_prods_tab4 luego de eliminar",MisMetodos.list_prods_tab4)
                index+=1
            return"Agregado correctamente"

    @staticmethod
    def validate_change_sale(list_data:list):
        if list_data[2] == "Fecha Venta":
            list_data[2]="fecha_venta"
            date = datetime.combine(list_data[3], datetime.min.time())
            date_year_before = datetime.now() - relativedelta(years=1)
            date_year_after = datetime.now() + relativedelta(years=1)
            if not (date_year_before <= date and date <= date_year_after):
                return "la fecha debe ser entre 1 año"
            else:
                GestionInventario.update_sale(list_data[0], list_data[2], list_data[3])
                return "Editado Correctamente"
        if list_data[2]=="Monto Total":
            list_data[2]="monto_total"
            if not MisMetodos.es_float(list_data[3]) or float(list_data[3]) < 1:
                return "Ingrese un precio valido"
            else:
                GestionInventario.update_sale(list_data[0], list_data[2], list_data[3])
                return "Editado Correctamente"
        if list_data[2]=="Cantidad":
            list_data[2]="cantidad"
            if not MisMetodos.es_int(list_data[3]) or int(list_data[3]) <= 0:
                return "Ingrese una cantidad valida"
            else:
                GestionInventario.update_sale_prod(list_data[0], list_data[1], list_data[2], list_data[3])
                return "Editado Correctamente"
        if list_data[2]=="Precio Unitario":
            list_data[2]="precio_unitario"
            if not MisMetodos.es_float(list_data[3]) or float(list_data[3]) <= 0:
                return "Ingrese un precio valido y una cantidad valida"
            else:
                GestionInventario.update_sale_prod(list_data[0], list_data[1], list_data[2], list_data[3])
                return "Editado Correctamente"
    @staticmethod
    def escribir_excel(data, nombre_archivo):
        columnas = ["Id", "nombre", "descripcion", "categoria", "precio", "stock"]
        wb = openpyxl.Workbook()
        ws = wb.active
        for col_num, nombre_columna in enumerate(columnas, start=1):
            ws.cell(row=1, column=col_num, value=nombre_columna)
        for row_num, row_data in enumerate(data, start=2):
            for col_num, col_value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_num, column=col_num, value=col_value)
                # Formato condicional para la columna 'stock'
                if col_num == 6 and col_value < 100:
                    fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    cell.fill = fill
        wb.save(nombre_archivo)




























